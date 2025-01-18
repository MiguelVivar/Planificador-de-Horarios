import os
import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def aplicar_bordes_y_relleno(cell, color_fondo):
    color_borde = "000000"  # Color del borde (negro)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(border_style="thin", color=color_borde),
        right=Side(border_style="thin", color=color_borde),
        top=Side(border_style="thin", color=color_borde),
        bottom=Side(border_style="thin", color=color_borde)
    )
    cell.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")

def guardar_en_excel(curso, ciclo, seccion, dia, empieza, termina, salon, tipo, profesor, verificar_conflicto):
    nombre_archivo = f"horario_{ciclo}_{seccion}.xlsx"
    ruta_carpeta = "data/horarios"
    
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)
    
    archivo_excel = os.path.join(ruta_carpeta, nombre_archivo)
    
    try:
        wb = openpyxl.load_workbook(archivo_excel)
        hoja = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.append(["Hora", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"])
    
    mensaje_conflicto, CONTINUAR = verificar_conflicto(hoja, curso, dia, empieza, termina, salon)
    if not CONTINUAR:
        return mensaje_conflicto
    
    horas = [
        ("07:45", "08:30"),
        ("08:30", "09:15"),
        ("09:15", "10:00"),
        ("10:00", "10:45"),
        ("10:45", "11:30"),
        ("11:30", "12:15"),
        ("12:15", "01:00"),
        ("01:00", "01:45")
    ]
    
    color_encabezado = "ADD8E6"
    color_fila_par = "F0F8FF"
    color_fila_impar = "FFFFFF"
    color_borde = "000000"

    for i, (hora_inicio, hora_fin) in enumerate(horas):
        hora_cell = hoja.cell(row=i + 2, column=1, value=f"{hora_inicio} - {hora_fin}")
        hora_cell.alignment = Alignment(horizontal="center", vertical="center")
        hora_cell.border = Border(
            left=Side(border_style="thin", color=color_borde),
            right=Side(border_style="thin", color=color_borde),
            top=Side(border_style="thin", color=color_borde),
            bottom=Side(border_style="thin", color=color_borde)
        )

    for col in range(1, 9):
        encabezado_cell = hoja.cell(row=1, column=col)
        encabezado_cell.fill = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type="solid")
        aplicar_bordes_y_relleno(encabezado_cell, color_encabezado)
    
    hora_inicio = datetime.strptime(empieza, "%H:%M")
    hora_termina = datetime.strptime(termina, "%H:%M")
    fila_inicio = (hora_inicio - datetime(2000, 1, 1, 7, 45)).seconds // 2700

    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    columna_dia = dias.index(dia) + 2
    
    for i in range(2, len(horas) + 2):
        for j in range(2, 9):
            cell = hoja.cell(row=i, column=j)
            if cell.value is None:
                cell.value = ""
                aplicar_bordes_y_relleno(cell, color_fila_par if i % 2 == 0 else color_fila_impar)

    for i in range(fila_inicio, len(horas)):
        if hora_inicio < hora_termina:
            cell = hoja.cell(row=i + 2, column=columna_dia)
            if cell.value:
                return f"Conflicto: Ya existe un curso en {dia} de {hora_inicio.strftime('%H:%M')} a {termina}"
            
            cell.value = f"{curso} ({salon}) - {tipo}"
            aplicar_bordes_y_relleno(cell, color_fila_par if i % 2 == 0 else color_fila_impar)
            hora_inicio += timedelta(minutes=45)
        else:
            break

    for col in range(1, 9):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in hoja[get_column_letter(col)])
        hoja.column_dimensions[get_column_letter(col)].width = max_length + 2

    wb.save(archivo_excel)
    return "Archivo guardado correctamente."
